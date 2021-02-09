import os
import json
from django.shortcuts import render, HttpResponse
from django.shortcuts import render_to_response
from .forms import f_callme, f_subscribe, f_email, f_product, f_contact, f_servicing
from django.core.mail import send_mail
from .models import product, material, category, spares, category_index_thumbs, gfaq, callme, subscribe, countries, states, cities, contact
from openpyxl import load_workbook
from django.core import serializers
from django.http import JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db import IntegrityError
script_dir = os.path.dirname(__file__) #<-- absolute dir the script is in

# Create your views here.
def v_index(request):
	if request.method == 'POST':
		if request.is_ajax:
			form = f_callme(request.POST)
			if form.is_valid():
				number = form.cleaned_data['number']
				mail = "yogibabanet@gmail.com"
				pwd = "@yogibabanet"
				value = send_mail("Hi", number, "Upendra", [mail], fail_silently=False, auth_user=mail, auth_password=pwd)
				return render(request, 'riplInd/index.html', {'title':'Send E-mail', 'success':'Congrats! The e-mail is received','form_call': form})		
	else:
		cat_index_thumbs = category_index_thumbs.objects.filter()
		form_call = f_callme()
		form_subscribe = f_subscribe()
	return render(request, 'riplInd/index.html', {'title':'Send E-mail','form_subscribe': form_subscribe,'form_call':form_call,'category':cat_index_thumbs})
		
def v_callme(request):
	if request.method == "POST":
		try:
			if 'callme' in request.POST:
				form = f_callme(request.POST)
				if form.is_valid():
					number = form.cleaned_data['number']
					obj = callme()
					obj.number = number
					obj.save()
					send_mail("hi", "ko", 'FirstBite@gmail.com', ["handa.anu@gmail.com"], fail_silently=False)
					return render(request,  'riplInd/response.html', {'success':"Our Customer Support Team will call you shortly."})
			elif 'subscribe' in request.POST:
				form = f_subscribe(request.POST)
				if form.is_valid():
					email = form.cleaned_data['email']
					obj = subscribe()
					obj.email = email
					obj.save()
				return render(request,  'riplInd/response.html', {'success':"You are now subscribed to our Newsletter."})
		except IntegrityError as e:
			return render(request,  'riplInd/response.html', {'success':e.message})
	else:
		form = f_callme()
		return render(request, 'riplInd/index.html', {'title':'Send E-mail','form': form, 'success':'it failed :('})
		
def v_about(request):
		return render(request, 'riplInd/about.html', {})
def v_product(request):
	if request.GET.get('id') > 1:
		id = id=request.GET.get('id')
	else:
		id = 1
	lo = product.objects.get(id=id)
	return render(request, 'riplInd/product.html', {'product':lo})
def v_project(request):
		return render(request, 'riplInd/project.html', {})
def v_contact(request):
	country = countries.objects.all()
	if request.method == "POST":
		form = f_contact(request.POST)
		try:
			if form.is_valid():
				reason = form.cleaned_data['reason']
				fname = form.cleaned_data['fname']
				lname = form.cleaned_data['lname']
				phone = form.cleaned_data['phone']
				email = form.cleaned_data['email']
				occupation = form.cleaned_data['occupation']
				country = request.POST['country']
				message = form.cleaned_data['message']
				obj = contact()
				obj.fname = fname
				obj.lname = lname
				obj.phone = phone
				obj.email = email
				obj.occupation = occupation
				obj.message = message
				obj.reason = reason
				obj.country = country
				obj.save()
				return render(request,'riplInd/contact.html', {'form':form, 'countries':country, 'success':"Thank you for writing to us. We Will Contact You Shortly"})
			else:
				return render(request,'riplInd/contact.html', {'form':form, 'countries':country, 'success':"Thanks. We Will Contact You lij"})
		except IntegrityError as e:
			return render(request,'riplInd/contact.html', {'form':form, 'countries':country, 'success':e.message})
	else:
		form = f_contact()
		return render(request,'riplInd/contact.html', {'form':form, 'countries':country})
def v_support(request):
	country = countries.objects.all()
	if request.method == "POST":
		form = f_servicing(request.POST)
		try:
			if form.is_valid():
				fname = form.cleaned_data['fname']
				lname = form.cleaned_data['lname']
				phone = form.cleaned_data['phone']
				email = form.cleaned_data['email']
				serial = form.cleaned_data['serial']
				country = request.POST['country']
				message = form.cleaned_data['message']
				obj = f_servicing()
				obj.fname = fname
				obj.lname = lname
				obj.phone = phone
				obj.email = email
				obj.serial = serial
				obj.message = message
				obj.country = country
				obj.save()
				return render(request,'riplInd/support.html', {'form':form, 'countries':country, 'success':"Thank you for writing to us. We Will Contact You Shortly"})
			else:
				return render(request,'riplInd/support.html', {'form':form, 'countries':country, 'success':"There was an error saving your information."})
		except IntegrityError as e:
			return render(request,'riplInd/support.html', {'form':form, 'countries':country, 'success':e.message})
	else:
		form = f_servicing()
		return render(request,'riplInd/support.html', {'form':form,'countries':country})
def v_faq(request):
		faq = gfaq.objects.all()
		return render(request,'riplInd/faq.html', {'faq':faq})
def v_fabric(request):
		return render(request,'riplInd/fabric.html', {})
def v_make(request):
		'''path = "static/spares.xlsx"
		file_path = os.path.join(script_dir, path)
		wb = load_workbook(filename=file_path, read_only=True)
		rows = wb['Sheet1'].iter_rows()'''
		spare = spares.objects.all()
		return render(request, 'riplInd/manufacture.html', {'spare':spare})
def v_collection(request):
	if request.GET.get('id')>1:
		id = request.GET.get('id')
	else:
		id = 1
	categories = category.objects.filter(flag=1)
	products = product.objects.filter(category__id='1', flag=1)
	return render(request, 'riplInd/collection.html', {'category':categories, 'items':products, 'id':id})
def v_itemfetch(request):
	if request.method == 'POST':
		if request.is_ajax():
			id = request.POST['id']
			lo = product.objects.filter(category__id=id)
			categories = category.objects.get(id=id)
			return render(request,  'riplInd/prolist.html', {'items':lo, 'fcategory':categories})